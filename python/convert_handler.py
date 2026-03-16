"""Document Conversion Handler - handles format conversions."""
import sys
import json
import os
import argparse
import re
import shutil
import io
import hashlib
from urllib.parse import urlparse
from pathlib import Path
from path_utils import (
    resolve_input_path,
    resolve_output_file,
    resolve_output_dir,
    default_single_output,
    default_first_input_output,
    create_unique_child_dir,
    resolve_chromium_executable,
    resolve_pandoc_executable,
    resolve_libreoffice_executable,
    resolve_poppler_bin_dir,
    resolve_tex_executable,
    resolve_wkhtmltopdf_executable,
)


def word_to_pdf(input_file: str, output_file: str = "") -> dict:
    """Convert Word document to PDF using LibreOffice (layout-preserving)."""
    try:
        input_file = resolve_input_path(input_file, __file__)
        if not os.path.exists(input_file):
            return {"success": False, "error": f"Input file not found: {input_file}. Please select the file using native file picker so absolute path is available."}
        output_file = resolve_output_file(output_file, __file__) if output_file and str(output_file).strip() else default_single_output(input_file, "", ".pdf")
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        return run_libreoffice_to_pdf(input_file, output_file)
    except Exception as e:
        return {"success": False, "error": str(e)}


def excel_to_pdf(input_file: str, output_file: str = "") -> dict:
    """Convert Excel workbook to PDF using LibreOffice (layout-preserving)."""

    try:
        input_file = resolve_input_path(input_file, __file__)
        if not os.path.exists(input_file):
            return {"success": False, "error": f"Input file not found: {input_file}. Please select the file using native file picker so absolute path is available."}
        output_file = resolve_output_file(output_file, __file__) if output_file and str(output_file).strip() else default_single_output(input_file, "", ".pdf")
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        return run_libreoffice_to_pdf(input_file, output_file)
    except Exception as e:
        return {"success": False, "error": str(e)}


def pptx_to_pdf(input_file: str, output_file: str = "") -> dict:
    """Convert PowerPoint to PDF.

    Default path uses LibreOffice headless export (keeps slide layout/design).
    Uses LibreOffice headless export to keep slide layout/design.
    """
    try:
        input_file = resolve_input_path(input_file, __file__)
        if not os.path.exists(input_file):
            return {"success": False, "error": f"Input file not found: {input_file}. Please select the file using native file picker so absolute path is available."}

        output_file = resolve_output_file(output_file, __file__) if output_file and str(output_file).strip() else default_single_output(input_file, "", ".pdf")
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        return run_libreoffice_to_pdf(input_file, output_file)
    except Exception as e:
        return {"success": False, "error": str(e)}


def run_libreoffice_to_pdf(input_file: str, output_file: str) -> dict:
    """Convert office document to PDF using LibreOffice headless export."""
    import subprocess

    soffice = resolve_libreoffice_executable()
    if not soffice:
        return {
            "success": False,
            "error": "libreoffice/soffice not found",
        }

    out_dir = str(Path(output_file).parent)
    expected_name = Path(input_file).with_suffix(".pdf").name
    generated_file = os.path.join(out_dir, expected_name)

    cmd = [
        soffice,
        "--headless",
        "--convert-to",
        "pdf",
        "--outdir",
        out_dir,
        input_file,
    ]

    result = subprocess.run(cmd, capture_output=True, text=True, timeout=240)
    if result.returncode != 0:
        detail = (result.stderr or result.stdout or "Unknown libreoffice error").strip()
        return {"success": False, "error": f"libreoffice conversion failed: {detail}"}

    if not os.path.exists(generated_file):
        return {"success": False, "error": f"libreoffice did not produce expected file: {generated_file}"}

    if os.path.abspath(generated_file) != os.path.abspath(output_file):
        shutil.move(generated_file, output_file)

    return {"success": True, "output": output_file, "engine": "libreoffice"}


def _convert_spreadsheet_to_xlsx_temp(input_file: str) -> tuple[str, str]:
    """Convert legacy spreadsheet to temporary .xlsx for openpyxl-compatible workflows."""
    import subprocess
    import tempfile

    soffice = resolve_libreoffice_executable()
    if not soffice:
        raise RuntimeError("libreoffice/soffice not found")

    temp_dir = tempfile.mkdtemp(prefix="excel_to_xlsx_")
    expected_name = Path(input_file).with_suffix(".xlsx").name
    output_file = os.path.join(temp_dir, expected_name)

    cmd = [
        soffice,
        "--headless",
        "--convert-to",
        "xlsx",
        "--outdir",
        temp_dir,
        input_file,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=240)
    if result.returncode != 0:
        detail = (result.stderr or result.stdout or "Unknown libreoffice error").strip()
        raise RuntimeError(f"libreoffice conversion failed: {detail}")

    if not os.path.exists(output_file):
        raise RuntimeError(f"libreoffice did not produce expected file: {output_file}")

    return output_file, temp_dir


def _guess_image_extension(image_name: str, data: bytes) -> str:
    """Infer image extension from name first, then magic bytes."""
    suffix = (Path(image_name).suffix or "").lower()
    if suffix in {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tif", ".tiff", ".jp2"}:
        return suffix

    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if data.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    if data.startswith((b"GIF87a", b"GIF89a")):
        return ".gif"
    if data.startswith(b"BM"):
        return ".bmp"
    if data.startswith(b"RIFF") and b"WEBP" in data[:16]:
        return ".webp"
    if data.startswith((b"II*\x00", b"MM\x00*")):
        return ".tiff"
    if data.startswith(b"\x00\x00\x00\x0cjP  \r\n\x87\n"):
        return ".jp2"

    return ".bin"


def _prepare_images_dir(output_file: str, images_dir_override: str = "") -> str:
    """Create a deterministic markdown-adjacent images directory."""
    md_path = Path(output_file)
    images_dir = Path(images_dir_override) if images_dir_override else (md_path.parent / f"{md_path.stem}_images")
    if images_dir.exists() and images_dir.is_dir():
        shutil.rmtree(images_dir)
    images_dir.mkdir(parents=True, exist_ok=True)
    return str(images_dir)


def _to_markdown_rel_path(target_path: str, markdown_file: str) -> str:
    """Build a markdown-friendly relative path from markdown file to target file."""
    base_dir = str(Path(markdown_file).parent)
    rel = os.path.relpath(target_path, base_dir).replace(os.sep, "/")
    if rel.startswith("."):
        return rel
    return f"./{rel}"


def _extract_pdf_images_by_page(
    input_file: str,
    markdown_file: str,
    extract_images: bool = True,
    images_dir_override: str = "",
) -> dict:
    """Extract embedded PDF images with pypdf and return page-indexed markdown paths."""
    if not extract_images:
        return {
            "images_by_page": {},
            "images_count": 0,
            "images_dir": "",
        }

    from pypdf import PdfReader

    reader = PdfReader(input_file)
    images_by_page = {}
    images_count = 0
    images_dir = ""
    if images_dir_override:
        images_dir = _prepare_images_dir(markdown_file, images_dir_override)
    hash_to_relpath = {}

    for page_index, page in enumerate(reader.pages, start=1):
        page_links = []
        page_images = getattr(page, "images", None)
        if not page_images:
            continue

        try:
            page_image_count = len(page_images)
        except Exception:
            continue

        for img_index in range(1, page_image_count + 1):
            try:
                image = page_images[img_index - 1]
                data = getattr(image, "data", None)
                if not data:
                    continue

                image_name = getattr(image, "name", f"image_{page_index}_{img_index}")
                ext = _guess_image_extension(str(image_name), data)
                digest = hashlib.sha256(data).hexdigest()

                if digest in hash_to_relpath:
                    page_links.append(hash_to_relpath[digest])
                    continue

                if not images_dir:
                    images_dir = _prepare_images_dir(markdown_file, images_dir_override)

                # Many system viewers cannot open JP2 reliably; convert to PNG when possible.
                convert_to_png = ext in {".jp2", ".tif", ".tiff", ".bin"}
                out_path = ""

                if convert_to_png:
                    try:
                        from PIL import Image

                        pil_img = getattr(image, "image", None)
                        if pil_img is None:
                            pil_img = Image.open(io.BytesIO(data))

                        file_name = f"page_{page_index:03d}_{img_index:02d}.png"
                        out_path = os.path.join(images_dir, file_name)
                        pil_img.save(out_path, format="PNG")
                    except Exception:
                        # If conversion is unavailable, keep raw stream as fallback.
                        file_name = f"page_{page_index:03d}_{img_index:02d}{ext}"
                        out_path = os.path.join(images_dir, file_name)
                        with open(out_path, "wb") as f:
                            f.write(data)
                else:
                    file_name = f"page_{page_index:03d}_{img_index:02d}{ext}"
                    out_path = os.path.join(images_dir, file_name)
                    with open(out_path, "wb") as f:
                        f.write(data)

                rel_path = _to_markdown_rel_path(out_path, markdown_file)
                hash_to_relpath[digest] = rel_path
                page_links.append(rel_path)
                images_count += 1
            except Exception:
                # Skip problematic image objects but keep overall conversion working.
                continue

        if page_links:
            images_by_page[page_index] = page_links

    return {
        "images_by_page": images_by_page,
        "images_count": images_count,
        "images_dir": images_dir,
    }


def _replace_glm_placeholders_with_pdf_crops(
    content: str,
    input_file: str,
    markdown_file: str,
    images_dir: str = "",
    dpi: int = 200,
    zero_based_pages: bool = False,
    page_offset: int = 0,
) -> tuple[str, int, str]:
    """Crop unresolved GLM placeholders from PDF page renders using page+bbox metadata."""
    pattern = re.compile(r"!\s*\[([^\]]*)\]\s*\(([^\)]*)\)")
    if not pattern.search(content):
        return content, 0, images_dir

    from pypdf import PdfReader

    try:
        from pdf2image import convert_from_path
    except Exception:
        return content, 0, images_dir

    total_pages = len(PdfReader(input_file).pages)
    page_cache = {}
    crop_cache = {}
    per_page_count = {}
    crop_count = 0

    def _resolve_page(raw_page: int) -> int | None:
        logical_page = (raw_page + 1 if zero_based_pages else raw_page) + page_offset
        candidate_pages = [logical_page]

        for p in candidate_pages:
            if 1 <= p <= total_pages:
                return p
        return None

    def _parse_bbox(target: str) -> tuple[float, float, float, float] | None:
        box_match = re.search(r"\b(?:bbox|rect|region)\b\s*[:=]\s*\[([^\]]+)\]", target, flags=re.IGNORECASE)
        if not box_match:
            return None
        nums = re.findall(r"-?\d+(?:\.\d+)?", box_match.group(1))
        if len(nums) < 4:
            return None

        x0, y0, x1, y1 = [float(v) for v in nums[:4]]
        left, right = (x0, x1) if x0 <= x1 else (x1, x0)
        top, bottom = (y0, y1) if y0 <= y1 else (y1, y0)
        return left, top, right, bottom

    def _get_page_image(page_no: int):
        if page_no not in page_cache:
            poppler_bin = resolve_poppler_bin_dir()
            kwargs = {
                "dpi": dpi,
                "first_page": page_no,
                "last_page": page_no,
            }
            if poppler_bin:
                kwargs["poppler_path"] = poppler_bin
            rendered = convert_from_path(
                input_file,
                **kwargs,
            )
            if not rendered:
                return None
            page_cache[page_no] = rendered[0]
        return page_cache.get(page_no)

    def repl(match: re.Match) -> str:
        nonlocal images_dir, crop_count

        alt = match.group(1)
        target = match.group(2).strip()

        if not re.search(r"\b(page|bbox|rect|region)\b", target, flags=re.IGNORECASE):
            return match.group(0)

        page_match = re.search(r"\bpage\b\s*[:=]\s*(\d+)", target, flags=re.IGNORECASE)
        if not page_match:
            return match.group(0)

        bbox = _parse_bbox(target)
        if bbox is None:
            return match.group(0)

        mapped_page = _resolve_page(int(page_match.group(1)))
        if mapped_page is None:
            return match.group(0)

        cache_key = (mapped_page, tuple(round(v, 3) for v in bbox))
        if cache_key in crop_cache:
            rel_path = crop_cache[cache_key]
            return f"![{alt}]({rel_path})"

        page_img = _get_page_image(mapped_page)
        if page_img is None:
            return match.group(0)

        width, height = page_img.size
        left, top, right, bottom = bbox

        # Support normalized bbox coordinates in [0, 1].
        if max(abs(left), abs(top), abs(right), abs(bottom)) <= 1.5:
            left *= width
            right *= width
            top *= height
            bottom *= height

        left = max(0.0, min(left, float(width)))
        right = max(0.0, min(right, float(width)))
        top = max(0.0, min(top, float(height)))
        bottom = max(0.0, min(bottom, float(height)))

        if right - left < 2 or bottom - top < 2:
            return match.group(0)

        if not images_dir:
            images_dir = _prepare_images_dir(markdown_file)

        per_page_count[mapped_page] = per_page_count.get(mapped_page, 0) + 1
        file_name = f"page_{mapped_page:03d}_crop_{per_page_count[mapped_page]:02d}.png"
        out_path = os.path.join(images_dir, file_name)

        try:
            crop = page_img.crop((int(left), int(top), int(right), int(bottom)))
            crop.save(out_path, format="PNG")
        except Exception:
            return match.group(0)

        rel_path = _to_markdown_rel_path(out_path, markdown_file)
        crop_cache[cache_key] = rel_path
        crop_count += 1
        return f"![{alt}]({rel_path})"

    replaced = pattern.sub(repl, content)
    return replaced, crop_count, images_dir


def _remove_glm_image_placeholders(content: str) -> str:
    """Remove unresolved GLM layout placeholders from markdown image syntax."""
    pattern = re.compile(r"!\s*\[[^\]]*\]\s*\(([^\)]*)\)")

    def repl(match: re.Match) -> str:
        target = match.group(1).strip()
        if re.search(r"\b(page|bbox|rect|region)\b", target, flags=re.IGNORECASE):
            return ""
        return match.group(0)

    cleaned = pattern.sub(repl, content)
    return re.sub(r"\n{3,}", "\n\n", cleaned).strip() + "\n"


def _md_escape(text: str) -> str:
    """Escape markdown-sensitive characters for table cells."""
    return (text or "").replace("|", r"\|").replace("\n", "<br>").strip()


def _rows_to_markdown_table(rows: list) -> str:
    """Convert 2D rows into a markdown table string."""
    if not rows:
        return ""

    normalized = []
    max_cols = 0
    for row in rows:
        cells = ["" if c is None else str(c).strip() for c in (row or [])]
        if any(cells):
            normalized.append(cells)
            max_cols = max(max_cols, len(cells))

    if not normalized or max_cols <= 0:
        return ""

    padded = []
    for row in normalized:
        if len(row) < max_cols:
            row = row + [""] * (max_cols - len(row))
        padded.append([_md_escape(cell) for cell in row])

    header = padded[0]
    body = padded[1:]

    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * max_cols) + " |",
    ]
    for row in body:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def _words_to_lines(words: list, y_tolerance: float = 3.0) -> list:
    """Cluster pdfplumber words into reading-order lines."""
    if not words:
        return []

    sorted_words = sorted(words, key=lambda w: (float(w.get("top", 0.0)), float(w.get("x0", 0.0))))
    lines = []
    current = []
    current_top = None

    for word in sorted_words:
        top = float(word.get("top", 0.0))
        if current_top is None or abs(top - current_top) <= y_tolerance:
            current.append(word)
            if current_top is None:
                current_top = top
            else:
                current_top = (current_top * (len(current) - 1) + top) / len(current)
        else:
            lines.append((current_top, current))
            current = [word]
            current_top = top

    if current:
        lines.append((current_top, current))

    result = []
    for line_top, line_words in lines:
        ordered = sorted(line_words, key=lambda w: float(w.get("x0", 0.0)))
        parts = []
        prev_x1 = None
        for w in ordered:
            txt = (w.get("text") or "").strip()
            if not txt:
                continue
            x0 = float(w.get("x0", 0.0))
            x1 = float(w.get("x1", x0))
            if prev_x1 is not None and x0 - prev_x1 > 2.5:
                parts.append(" ")
            parts.append(txt)
            prev_x1 = x1
        line_text = "".join(parts).strip()
        if line_text:
            result.append({"top": line_top, "text": line_text})
    return result


def _extract_pdf_markdown_with_pdfplumber(input_file: str, images_by_page: dict) -> tuple[str, int]:
    """Extract markdown from PDF using pdfplumber-like layout methods."""
    import pdfplumber

    pages_md = []
    rendered_pages = 0
    seen_global = set()

    with pdfplumber.open(input_file) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            items = []

            table_blocks = []
            try:
                tables = page.find_tables(table_settings={
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "snap_tolerance": 3,
                    "join_tolerance": 3,
                    "intersection_tolerance": 3,
                })
            except Exception:
                tables = []

            for t in tables:
                try:
                    rows = t.extract() or []
                    table_md = _rows_to_markdown_table(rows)
                    if not table_md:
                        continue
                    top = float(t.bbox[1]) if getattr(t, "bbox", None) else 0.0
                    bottom = float(t.bbox[3]) if getattr(t, "bbox", None) else top
                    table_blocks.append({"top": top, "bottom": bottom, "content": table_md})
                    items.append({"top": top, "kind": "table", "content": table_md})
                except Exception:
                    continue

            try:
                words = page.extract_words(
                    x_tolerance=2,
                    y_tolerance=2,
                    use_text_flow=True,
                    keep_blank_chars=False,
                )
            except Exception:
                words = []

            lines = _words_to_lines(words, y_tolerance=3.0)
            for line in lines:
                line_top = float(line["top"])
                # Skip text likely inside already extracted table regions.
                if any(tb["top"] - 2 <= line_top <= tb["bottom"] + 2 for tb in table_blocks):
                    continue
                items.append({"top": line_top, "kind": "text", "content": line["text"]})

            page_images = images_by_page.get(page_index, [])
            unique_page_images = []
            for rel_path in page_images:
                if rel_path in seen_global:
                    continue
                seen_global.add(rel_path)
                unique_page_images.append(rel_path)

            if not items and not unique_page_images:
                continue

            rendered_pages += 1
            page_parts = []
            for item in sorted(items, key=lambda it: (it["top"], 0 if it["kind"] == "text" else 1)):
                if item["kind"] == "table":
                    page_parts.append(item["content"] + "\n")
                else:
                    page_parts.append(item["content"] + "\n")

            for img_idx, rel_path in enumerate(unique_page_images, start=1):
                page_parts.append(f"![Image {img_idx}]({rel_path})\n")

            pages_md.append("\n".join(page_parts).strip() + "\n")

    return "\n---\n\n".join(pages_md), rendered_pages


def pdf_to_markdown(input_file: str, output_file: str = "",
                    provider: str = "auto", credentials: dict = None,
                    lang: str = "chi_sim+eng", extract_images: bool = True,
                    images_dir_override: str = "") -> dict:
    """Convert PDF to Markdown. Prefer GLM-OCR markdown when API key is available."""
    try:
        input_file = resolve_input_path(input_file, __file__)
        if not os.path.exists(input_file):
            return {"success": False, "error": f"Input file not found: {input_file}. Please select the file using native file picker so absolute path is available."}
        src = Path(input_file).resolve()
        if output_file and str(output_file).strip():
            output_file = resolve_output_file(output_file, __file__)
        elif extract_images:
            output_dir = src.parent / src.stem
            output_dir.mkdir(parents=True, exist_ok=True)
            output_file = str(output_dir / f"{src.stem}.md")
        else:
            output_file = str(src.with_suffix(".md"))

        if extract_images and not images_dir_override:
            images_dir_override = str(Path(output_file).parent / "img")

        Path(output_file).parent.mkdir(parents=True, exist_ok=True)

        credentials = credentials or {}
        api_key = credentials.get("api_key") if isinstance(credentials, dict) else None
        use_glm = provider in ("auto", "glm") and bool(api_key)

        images_by_page = {}
        images_count = 0
        images_dir = ""
        if use_glm and extract_images:
            images_dir = _prepare_images_dir(output_file, images_dir_override)
        if not use_glm:
            image_meta = _extract_pdf_images_by_page(
                input_file,
                output_file,
                extract_images=extract_images,
                images_dir_override=images_dir_override,
            )
            images_by_page = image_meta["images_by_page"]
            images_count = image_meta["images_count"]
            images_dir = image_meta["images_dir"]

        # Reuse OCR module's GLM pipeline when key is present; it already returns markdown text.
        if use_glm:
            from ocr_handler import ocr_pdf
            glm_res = ocr_pdf(
                pdf_path=input_file,
                lang=lang,
                provider="glm",
                credentials={"api_key": api_key},
            )
            if glm_res.get("success") and (glm_res.get("text") or "").strip():
                content = glm_res.get("text", "")
                crop_count = 0
                if extract_images:
                    page_texts = glm_res.get("page_texts") if isinstance(glm_res.get("page_texts"), list) else []
                    use_segment_mode = len(page_texts) > 0

                    if use_segment_mode:
                        out_parts = []
                        for idx, part in enumerate(page_texts):
                            zero_based_pages = bool(re.search(r"\bpage\b\s*[:=]\s*0\b", part, flags=re.IGNORECASE))
                            part, part_crop_count, images_dir = _replace_glm_placeholders_with_pdf_crops(
                                part,
                                input_file,
                                output_file,
                                images_dir=images_dir,
                                zero_based_pages=zero_based_pages,
                                page_offset=idx,
                            )
                            crop_count += part_crop_count
                            out_parts.append(part)
                        content = "\n\n---\n\n".join(out_parts)
                    else:
                        zero_based_pages = bool(re.search(r"\bpage\b\s*[:=]\s*0\b", content, flags=re.IGNORECASE))
                        content, crop_count, images_dir = _replace_glm_placeholders_with_pdf_crops(
                            content,
                            input_file,
                            output_file,
                            images_dir=images_dir,
                            zero_based_pages=zero_based_pages,
                        )
                # Always remove placeholders that were not replaced.
                content = _remove_glm_image_placeholders(content)
                with open(output_file, "w", encoding="utf-8") as f:
                    f.write(content)
                return {
                    "success": True,
                    "output": output_file,
                    "source": "glm",
                    "pages": glm_res.get("pages"),
                    "images_dir": images_dir,
                    "images_count": crop_count,
                }
            if provider == "glm":
                return {
                    "success": False,
                    "error": glm_res.get("error", "GLM OCR failed"),
                    "source": "glm",
                    "images_dir": images_dir,
                    "images_count": images_count,
                }

        try:
            content, parsed_pages = _extract_pdf_markdown_with_pdfplumber(input_file, images_by_page)
            source_name = "pdfplumber"
        except Exception:
            # Safety fallback when pdfplumber is unavailable or a PDF is malformed.
            from pypdf import PdfReader
            reader = PdfReader(input_file)
            lines = []
            seen_global = set()
            for i, page in enumerate(reader.pages):
                text = page.extract_text() or ""
                page_no = i + 1
                page_images = images_by_page.get(page_no, [])

                if not text.strip() and not page_images:
                    continue

                page_parts = []
                if text.strip():
                    page_parts.append(text.strip() + "\n")
                unique_page_images = []
                for rel_path in page_images:
                    if rel_path in seen_global:
                        continue
                    seen_global.add(rel_path)
                    unique_page_images.append(rel_path)

                for img_idx, rel_path in enumerate(unique_page_images, start=1):
                    page_parts.append(f"![Image {img_idx}]({rel_path})\n")

                lines.append("\n".join(page_parts).strip() + "\n")

            content = "\n---\n\n".join(lines)
            parsed_pages = len(lines)
            source_name = "pypdf"

        if not content.strip() and images_count > 0:
            # Keep markdown non-empty when the PDF has only images.
            img_lines = ["## Images\n"]
            flat = []
            for page_no in sorted(images_by_page.keys()):
                for rel_path in images_by_page.get(page_no, []):
                    flat.append((page_no, rel_path))
            for idx, (_page_no, rel_path) in enumerate(flat, start=1):
                img_lines.append(f"![Image {idx}]({rel_path})")
            content = "\n".join(img_lines).strip() + "\n"

        with open(output_file, "w", encoding="utf-8") as f:
            f.write(content)
        return {
            "success": True,
            "output": output_file,
            "source": source_name,
            "pages": parsed_pages,
            "images_dir": images_dir,
            "images_count": images_count,
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def word_to_markdown(input_file: str, output_file: str = "",
                     provider: str = "auto", credentials: dict = None,
                     lang: str = "chi_sim+eng", extract_images: bool = True) -> dict:
    """Convert Word to Markdown by chaining Word->PDF and PDF->Markdown."""
    try:
        input_file = resolve_input_path(input_file, __file__)
        if not os.path.exists(input_file):
            return {
                "success": False,
                "error": f"Input file not found: {input_file}. Please select the file using native file picker so absolute path is available."
            }

        src = Path(input_file).resolve()
        if output_file and str(output_file).strip():
            resolved_output = resolve_output_file(output_file, __file__)
        elif extract_images:
            output_dir = src.parent / src.stem
            output_dir.mkdir(parents=True, exist_ok=True)
            resolved_output = str(output_dir / f"{src.stem}.md")
        else:
            resolved_output = str(src.with_suffix(".md"))

        images_dir_override = ""
        if extract_images:
            images_dir_override = str(Path(resolved_output).parent / "img")

        # Use an adjacent temp PDF so we can reuse existing PDF->Markdown pipeline.
        temp_pdf = default_single_output(input_file, "_tmp_word_md", ".pdf")
        Path(temp_pdf).parent.mkdir(parents=True, exist_ok=True)

        pdf_res = word_to_pdf(input_file, temp_pdf)
        if not pdf_res.get("success"):
            return pdf_res

        try:
            md_res = pdf_to_markdown(
                input_file=pdf_res.get("output", temp_pdf),
                output_file=resolved_output,
                provider=provider,
                credentials=credentials,
                lang=lang,
                extract_images=extract_images,
                images_dir_override=images_dir_override,
            )
            if md_res.get("success"):
                md_res["source"] = f"word_to_pdf+{md_res.get('source', 'pdf')}"
                md_res["pipeline"] = "word->pdf->markdown"
            return md_res
        finally:
            if os.path.exists(temp_pdf):
                os.remove(temp_pdf)
    except Exception as e:
        return {"success": False, "error": str(e)}


def images_to_pdf(input_files: list, output_file: str = "") -> dict:
    """Convert one or more images to PDF using Pandoc via temporary HTML."""
    import tempfile
    try:
        if not input_files:
            return {"success": False, "error": "No input files provided"}
        input_files = [resolve_input_path(p, __file__) for p in input_files]
        for p in input_files:
            if not os.path.exists(p):
                return {"success": False, "error": f"Input file not found: {p}. Please select files using native file picker so absolute paths are available."}
        output_file = resolve_output_file(output_file, __file__) if output_file and str(output_file).strip() else default_first_input_output(input_files, "_images", ".pdf", __file__)
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)

        html_parts = [
            "<!doctype html><html><head><meta charset='utf-8'><style>",
            "@page { size: A4; margin: 20mm; }",
            "body { font-family: sans-serif; margin: 0; }",
            ".page { page-break-after: always; text-align: center; }",
            ".page:last-child { page-break-after: auto; }",
            "img { max-width: 100%; max-height: 250mm; object-fit: contain; }",
            "</style></head><body>",
        ]
        for img_path in input_files:
            uri = Path(img_path).resolve().as_uri()
            html_parts.append(f"<div class='page'><img src='{uri}'></div>")
        html_parts.append("</body></html>")
        html_doc = "".join(html_parts)

        with tempfile.NamedTemporaryFile("w", suffix=".html", delete=False, encoding="utf-8") as tmp:
            tmp.write(html_doc)
            tmp_path = tmp.name

        try:
            return run_pandoc_to_pdf(tmp_path, output_file, from_format="html", extra_args=["--standalone"])
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
    except Exception as e:
        return {"success": False, "error": str(e)}


def html_to_pdf(input_file: str, output_file: str = "") -> dict:
    """Convert HTML file to PDF by rendering the original page."""
    input_file = resolve_input_path(input_file, __file__)
    if not os.path.exists(input_file):
        return {"success": False, "error": f"Input file not found: {input_file}. Please select the file using native file picker so absolute path is available."}
    output_file = resolve_output_file(output_file, __file__) if output_file and str(output_file).strip() else default_single_output(input_file, "", ".pdf")
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)
    return run_html_renderer_to_pdf(input_file, output_file)


def run_html_renderer_to_pdf(input_file: str, output_file: str) -> dict:
    """Render HTML directly to PDF (Chromium/Chrome first, then wkhtmltopdf, then weasyprint)."""
    import subprocess

    errors = []

    # 1) Browser-grade rendering: Chromium/Chrome headless print-to-pdf.
    browser = resolve_chromium_executable()
    if browser:
        try:
            html_uri = Path(input_file).resolve().as_uri()
            cmd = [
                browser,
                "--headless=new",
                "--disable-gpu",
                "--no-sandbox",
                "--allow-file-access-from-files",
                "--virtual-time-budget=5000",
                f"--print-to-pdf={output_file}",
                html_uri,
            ]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=240)
            if result.returncode == 0 and os.path.exists(output_file):
                return {"success": True, "output": output_file, "engine": Path(browser).name}
            errors.append("[chromium] " + (result.stderr or result.stdout or "Unknown error").strip())
        except Exception as e:
            errors.append(f"[chromium] {e}")
    else:
        errors.append("[chromium] not found")

    # 2) wkhtmltopdf fallback.
    wkhtml = resolve_wkhtmltopdf_executable()

    if wkhtml:
        cmd = [
            wkhtml,
            "--enable-local-file-access",
            "--javascript-delay",
            "1200",
            "--encoding",
            "utf-8",
            input_file,
            output_file,
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
        if result.returncode == 0 and os.path.exists(output_file):
            return {"success": True, "output": output_file, "engine": "wkhtmltopdf"}
        errors.append("[wkhtmltopdf] " + (result.stderr or result.stdout or "Unknown error").strip())
    else:
        errors.append("[wkhtmltopdf] not found")

    # 3) weasyprint fallback.
    try:
        import weasyprint

        weasyprint.HTML(filename=input_file).write_pdf(output_file)
        if os.path.exists(output_file):
            return {"success": True, "output": output_file, "engine": "weasyprint"}
        errors.append("[weasyprint] failed without output")
    except Exception as e:
        errors.append(f"[weasyprint] {e}")

    return {
        "success": False,
        "error": "HTML rendering to PDF failed. " + " | ".join(errors),
    }


def normalize_markdown_latex(md_text: str) -> str:
    """Normalize OCR-noisy LaTeX in markdown while preserving math delimiters."""

    def normalize_ocr_latex(expr: str) -> str:
        s = (expr or "").strip()
        if not s:
            return s

        # Normalize common OCR spacing around LaTeX control sequences.
        s = re.sub(r"\\([a-zA-Z]+)\s+\{", r"\\\1{", s)
        s = re.sub(r"([_^])\s*\{\s*", r"\1{", s)
        s = re.sub(r"\s*\}", "}", s)

        def collapse_group_content(m: re.Match) -> str:
            cmd = m.group(1)
            content = m.group(2)
            content = re.sub(r"(?<=\d)\s+(?=\d)", "", content)
            content = re.sub(r"(?<=[A-Za-z])\s+(?=[A-Za-z])", "", content)
            content = re.sub(r"(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fff])", "", content)
            return f"\\{cmd}{{{content}}}"

        # OCR often inserts spaces in these braced argument commands.
        for cmd in ["mathrm", "text", "operatorname", "mathbf", "mathit"]:
            pattern = re.compile(rf"\\({cmd})\{{([^{{}}]*)\}}")
            s = pattern.sub(collapse_group_content, s)

        # Global digit compaction handles artifacts like '2 0' -> '20'.
        s = re.sub(r"(?<=\d)\s+(?=\d)", "", s)
        return s

    def repl_block(m: re.Match) -> str:
        return f"$${normalize_ocr_latex(m.group(1))}$$"

    def repl_inline(m: re.Match) -> str:
        return f"${normalize_ocr_latex(m.group(1))}$"

    # Block math first: $$ ... $$
    block_pattern = re.compile(r"(?<!\\)\$\$(.+?)(?<!\\)\$\$", re.DOTALL)
    processed = block_pattern.sub(repl_block, md_text)

    # Inline math: $ ... $ (avoid matching $$ ... $$ and escaped \$)
    inline_pattern = re.compile(r"(?<!\\)\$(?!\$)(.+?)(?<!\\)\$(?!\$)")
    processed = inline_pattern.sub(repl_inline, processed)

    return processed

def build_markdown_pdf_header_tex() -> str:
    """Build a minimal LaTeX header for chemistry-friendly PDF layout."""
    return r"""
\usepackage{fontspec}
\IfFontExistsTF{Noto Serif CJK SC}{\setmainfont{Noto Serif CJK SC}}{\setmainfont{DejaVu Serif}}
\usepackage{amsmath,amssymb}
\usepackage{mathtools}
\setlength{\parindent}{0pt}
\setlength{\parskip}{0.35em}
\allowdisplaybreaks
\setlength{\abovedisplayskip}{8pt}
\setlength{\belowdisplayskip}{8pt}
""".strip() + "\n"


def _infer_remote_image_extension(url: str, content_type: str = "") -> str:
    """Infer file extension for a downloaded remote image."""
    url_path = urlparse(url).path or ""
    suffix = Path(url_path).suffix.lower()
    if suffix in {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".svg", ".tif", ".tiff", ".jp2", ".avif"}:
        return suffix

    ctype = (content_type or "").split(";", 1)[0].strip().lower()
    by_type = {
        "image/png": ".png",
        "image/jpeg": ".jpg",
        "image/gif": ".gif",
        "image/bmp": ".bmp",
        "image/webp": ".webp",
        "image/svg+xml": ".svg",
        "image/tiff": ".tiff",
        "image/jp2": ".jp2",
        "image/avif": ".avif",
    }
    return by_type.get(ctype, ".img")


def _rewrite_markdown_remote_images(md_text: str, remote_dir: str) -> tuple[str, int, list[str]]:
    """Download remote markdown images and rewrite links to local temporary files."""
    import requests

    Path(remote_dir).mkdir(parents=True, exist_ok=True)

    errors = []
    replaced = 0
    url_to_local = {}

    pattern = re.compile(r"!\[([^\]]*)\]\(([^\)]+)\)")

    def parse_target(raw_target: str) -> tuple[str, str]:
        target = (raw_target or "").strip()
        if not target:
            return "", ""
        if target.startswith("<") and ">" in target:
            end = target.find(">")
            return target[1:end].strip(), target[end + 1:].strip()
        if " " in target:
            first, rest = target.split(" ", 1)
            return first.strip(), rest.strip()
        return target, ""

    def repl(match: re.Match) -> str:
        nonlocal replaced
        alt = match.group(1)
        raw_target = match.group(2)
        url, trailing = parse_target(raw_target)

        if not re.match(r"^https?://", url, flags=re.IGNORECASE):
            return match.group(0)

        try:
            if url not in url_to_local:
                resp = requests.get(
                    url,
                    timeout=20,
                    headers={
                        "User-Agent": "Mozilla/5.0 (FormatFlex Markdown->PDF)",
                    },
                )
                resp.raise_for_status()

                ext = _infer_remote_image_extension(url, resp.headers.get("Content-Type", ""))
                digest = hashlib.sha256(resp.content).hexdigest()[:12]
                local_name = f"remote_{len(url_to_local) + 1:03d}_{digest}{ext}"
                local_path = Path(remote_dir) / local_name
                with open(local_path, "wb") as f:
                    f.write(resp.content)
                url_to_local[url] = f"./{Path(remote_dir).name}/{local_name}"

            local_ref = url_to_local[url]
            replaced += 1
            if trailing:
                return f"![{alt}]({local_ref} {trailing})"
            return f"![{alt}]({local_ref})"
        except Exception as e:
            errors.append(f"{url}: {e}")
            return match.group(0)

    rewritten = pattern.sub(repl, md_text)
    return rewritten, replaced, errors


def markdown_to_pdf(input_file: str, output_file: str = "") -> dict:
    """Convert Markdown file to rendered PDF using Pandoc."""
    import tempfile

    try:
        input_file = resolve_input_path(input_file, __file__)
        if not os.path.exists(input_file):
            return {
                "success": False,
                "error": f"Input file not found: {input_file}. Please select the file using native file picker so absolute path is available."
            }

        # Compute output path from the original markdown file, not from temp HTML path.
        output_file = (
            resolve_output_file(output_file, __file__)
            if output_file and str(output_file).strip()
            else default_single_output(input_file, "", ".pdf")
        )
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)

        with open(input_file, "r", encoding="utf-8") as f:
            md_text = f.read()

        md_text = normalize_markdown_latex(md_text)

        source_dir = str(Path(input_file).resolve().parent)

        with tempfile.TemporaryDirectory(prefix="fmtflex_md_pdf_") as temp_dir:
            header_path = str(Path(temp_dir) / "header.tex")
            tmp_path = str(Path(temp_dir) / "input.md")
            remote_dir = str(Path(temp_dir) / "remote_assets")

            md_text, _replaced_count, _download_errors = _rewrite_markdown_remote_images(md_text, remote_dir)

            with open(header_path, "w", encoding="utf-8") as header:
                header.write(build_markdown_pdf_header_tex())

            with open(tmp_path, "w", encoding="utf-8") as tmp:
                tmp.write(md_text)

            resource_path = os.pathsep.join([temp_dir, source_dir])

            result = run_pandoc_to_pdf(
                tmp_path,
                output_file,
                from_format="markdown+tex_math_dollars",
                extra_args=[
                    "--standalone",
                    "--include-in-header",
                    header_path,
                    "--resource-path",
                    resource_path,
                ],
            )

            if result.get("success"):
                return result

            # Helpful diagnostics when remote image download failed but conversion also failed.
            if _download_errors:
                err = result.get("error", "")
                result["error"] = err + " | Remote image download warnings: " + " ; ".join(_download_errors[:5])
            return result
    except Exception as e:
        return {"success": False, "error": str(e)}


def run_pandoc_to_pdf(
    input_file: str,
    output_file: str,
    from_format: str | None = None,
    extra_args: list[str] | None = None,
    preferred_engines: list[str] | None = None,
) -> dict:
    """Run Pandoc to generate a PDF, trying available engines in priority order."""
    import subprocess

    def resolve_pdf_engine(engine: str) -> str | None:
        if engine == "wkhtmltopdf":
            return resolve_wkhtmltopdf_executable()
        if engine in ("xelatex", "pdflatex"):
            return resolve_tex_executable(engine)
        return shutil.which(engine)

    pandoc_bin = resolve_pandoc_executable()
    if not pandoc_bin:
        return {
            "success": False,
            "error": "pandoc not found. Install with: sudo apt install pandoc",
        }

    default_order = ["xelatex", "pdflatex", "weasyprint", "wkhtmltopdf"]
    engine_order = preferred_engines if preferred_engines else default_order
    available_engines = [(engine, resolve_pdf_engine(engine)) for engine in engine_order]
    available_engines = [(engine, engine_bin) for engine, engine_bin in available_engines if engine_bin]
    if not available_engines:
        return {
            "success": False,
            "error": "No PDF engine found for pandoc. Install one of: weasyprint, wkhtmltopdf, xelatex, pdflatex",
        }

    base_cmd = [pandoc_bin, input_file, "-o", output_file]
    if from_format:
        base_cmd += ["-f", from_format]
    if extra_args:
        base_cmd += extra_args

    errors = []
    for engine, engine_bin in available_engines:
        cmd = base_cmd + ["--pdf-engine", engine_bin]
        # HTML-based engines usually need MathML output for math formula rendering.
        if engine in ("weasyprint", "wkhtmltopdf") and "--mathml" not in cmd:
            cmd += ["--mathml"]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if result.returncode == 0 and os.path.exists(output_file):
            return {"success": True, "output": output_file, "engine": engine}
        detail = (result.stderr or result.stdout or "Unknown pandoc error").strip()
        errors.append(f"[{engine}] {detail}")

    return {
        "success": False,
        "error": "Pandoc PDF conversion failed. " + " | ".join(errors),
    }


def docx_to_txt(input_file: str, output_file: str = "") -> dict:
    """Extract text from Word document."""
    from docx import Document
    try:
        input_file = resolve_input_path(input_file, __file__)
        if not os.path.exists(input_file):
            return {"success": False, "error": f"Input file not found: {input_file}. Please select the file using native file picker so absolute path is available."}
        output_file = resolve_output_file(output_file, __file__) if output_file and str(output_file).strip() else default_single_output(input_file, "", ".txt")
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        doc = Document(input_file)
        text = "\n".join([p.text for p in doc.paragraphs])
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(text)
        return {"success": True, "output": output_file}
    except Exception as e:
        return {"success": False, "error": str(e)}


def _to_markdown_table(rows: list[list[str]]) -> str:
    """Build a Markdown table from row values using the first row as header."""
    if not rows:
        return ""

    col_count = max(len(r) for r in rows) if rows else 0
    if col_count == 0:
        return ""

    normalized = []
    for row in rows:
        normalized.append(row + [""] * (col_count - len(row)))

    def esc(cell: str) -> str:
        return str(cell).replace("|", "\\|").replace("\n", " ").replace("\r", " ").strip()

    header = normalized[0]
    body = normalized[1:] if len(normalized) > 1 else []

    lines = [
        "| " + " | ".join(esc(c) for c in header) + " |",
        "| " + " | ".join("---" for _ in range(col_count)) + " |",
    ]
    for row in body:
        lines.append("| " + " | ".join(esc(c) for c in row) + " |")

    return "\n".join(lines)


def excel_to_markdown(input_file: str, output_dir: str = "") -> dict:
    """Convert each Excel sheet to a Markdown file with a table."""
    import openpyxl

    try:
        input_file = resolve_input_path(input_file, __file__)
        if not os.path.exists(input_file):
            return {
                "success": False,
                "error": f"Input file not found: {input_file}. Please select the file using native file picker so absolute path is available."
            }

        supported_exts = {".xlsx", ".xlsm", ".xltx", ".xltm"}
        ext = Path(input_file).suffix.lower()
        workbook_input = input_file
        temp_dir = ""

        if ext in supported_exts:
            pass
        elif ext in {".xls", ".ods"}:
            workbook_input, temp_dir = _convert_spreadsheet_to_xlsx_temp(input_file)
        else:
            return {
                "success": False,
                "error": "Unsupported Excel format for this operation. Supported: .xlsx, .xlsm, .xltx, .xltm, .xls, .ods"
            }

        if output_dir and str(output_dir).strip():
            output_dir = resolve_output_dir(output_dir, __file__)
            Path(output_dir).mkdir(parents=True, exist_ok=True)
        else:
            output_dir = create_unique_child_dir(input_file, "markdown")

        wb = openpyxl.load_workbook(workbook_input, data_only=True)
        outputs = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(v) if v is not None else "" for v in row])

            safe_name = "".join(c if c.isalnum() else "_" for c in sheet_name) or "Sheet"
            out_path = os.path.join(output_dir, f"{safe_name}.md")

            title = f"# {sheet_name}\n\n"
            table = _to_markdown_table(rows)
            content = title + (table if table else "(empty sheet)") + "\n"

            with open(out_path, "w", encoding="utf-8") as f:
                f.write(content)
            outputs.append(out_path)

        return {"success": True, "outputs": outputs, "output": output_dir}
    except Exception as e:
        return {"success": False, "error": str(e)}
    finally:
        if 'temp_dir' in locals() and temp_dir and os.path.isdir(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)


def excel_to_csv(input_file: str, output_dir: str = "") -> dict:
    """Convert Excel sheets to CSV files."""
    import openpyxl
    try:
        input_file = resolve_input_path(input_file, __file__)
        if not os.path.exists(input_file):
            return {"success": False, "error": f"Input file not found: {input_file}. Please select the file using native file picker so absolute path is available."}

        supported_exts = {".xlsx", ".xlsm", ".xltx", ".xltm"}
        ext = Path(input_file).suffix.lower()
        workbook_input = input_file
        temp_dir = ""

        if ext in supported_exts:
            pass
        elif ext in {".xls", ".ods"}:
            workbook_input, temp_dir = _convert_spreadsheet_to_xlsx_temp(input_file)
        else:
            return {
                "success": False,
                "error": "Unsupported Excel format for this operation. Supported: .xlsx, .xlsm, .xltx, .xltm, .xls, .ods"
            }

        if output_dir and str(output_dir).strip():
            output_dir = resolve_output_dir(output_dir, __file__)
            Path(output_dir).mkdir(parents=True, exist_ok=True)
        else:
            output_dir = create_unique_child_dir(input_file, "csv")
        wb = openpyxl.load_workbook(workbook_input, data_only=True)
        os.makedirs(output_dir, exist_ok=True)
        outputs = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            safe_name = "".join(c if c.isalnum() else "_" for c in sheet_name)
            out_path = os.path.join(output_dir, f"{safe_name}.csv")
            import csv
            with open(out_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow([str(v) if v is not None else "" for v in row])
            outputs.append(out_path)
        return {"success": True, "outputs": outputs, "output": output_dir}
    except Exception as e:
        return {"success": False, "error": str(e)}
    finally:
        if 'temp_dir' in locals() and temp_dir and os.path.isdir(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)


def _sanitize_sheet_title(title: str, fallback: str, used_titles: set[str]) -> str:
    """Normalize an Excel sheet title and keep it unique."""
    safe = re.sub(r'[\\/*?:\[\]]', "_", (title or "").strip())
    safe = safe[:31].strip() or fallback
    candidate = safe
    suffix = 1
    while candidate in used_titles:
        suffix_text = f"_{suffix}"
        candidate = (safe[: 31 - len(suffix_text)] or fallback[: 31 - len(suffix_text)] or "Sheet") + suffix_text
        suffix += 1
    used_titles.add(candidate)
    return candidate


def _parse_markdown_table_row(line: str) -> list[str]:
    """Parse a pipe-style markdown table row into cells."""
    stripped = line.strip()
    if stripped.startswith("|"):
        stripped = stripped[1:]
    if stripped.endswith("|"):
        stripped = stripped[:-1]
    cells = re.split(r'(?<!\\)\|', stripped)
    return [cell.replace(r"\|", "|").strip() for cell in cells]


def _is_markdown_alignment_row(cells: list[str]) -> bool:
    """Detect markdown alignment separator rows like | --- | :---: |."""
    if not cells:
        return False
    return all(re.fullmatch(r":?-{3,}:?", cell or "") for cell in cells)


def _extract_markdown_tables(md_text: str) -> list[dict]:
    """Extract pipe-style markdown tables and nearby heading titles."""
    lines = md_text.splitlines()
    tables = []
    last_heading = "Sheet"
    index = 0

    while index < len(lines):
        line = lines[index].rstrip()
        stripped = line.strip()

        heading_match = re.match(r"^(#{1,6})\s+(.*)$", stripped)
        if heading_match:
            last_heading = heading_match.group(2).strip() or "Sheet"
            index += 1
            continue

        if not stripped.startswith("|"):
            index += 1
            continue

        block = []
        while index < len(lines) and lines[index].strip().startswith("|"):
            block.append(lines[index].rstrip())
            index += 1

        if len(block) < 2:
            continue

        header = _parse_markdown_table_row(block[0])
        align = _parse_markdown_table_row(block[1])
        if not header or not _is_markdown_alignment_row(align):
            continue

        rows = [header]
        for raw_row in block[2:]:
            row = _parse_markdown_table_row(raw_row)
            if any(cell.strip() for cell in row):
                rows.append(row)

        if rows:
            tables.append({
                "title": last_heading or f"Table_{len(tables) + 1}",
                "rows": rows,
            })

    return tables


def _write_excel_workbook(output_file: str, sheets: list[dict]) -> dict:
    """Write sheet data to an Excel workbook."""
    import openpyxl

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    used_titles = set()
    total_rows = 0

    for index, sheet in enumerate(sheets, start=1):
        rows = sheet.get("rows") or []
        if not rows:
            continue

        title = _sanitize_sheet_title(sheet.get("title", ""), f"Sheet_{index:02d}", used_titles)
        ws = wb.create_sheet(title=title)

        max_cols = max((len(row) for row in rows), default=0)
        for row_index, row in enumerate(rows, start=1):
            padded = list(row) + [""] * (max_cols - len(row))
            for col_index, cell in enumerate(padded, start=1):
                ws.cell(row=row_index, column=col_index, value="" if cell is None else str(cell))

        total_rows += len(rows)

    if not wb.sheetnames:
        ws = wb.create_sheet(title="Sheet_01")
        ws.cell(row=1, column=1, value="(empty)")

    Path(output_file).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    return {
        "success": True,
        "output": output_file,
        "sheets": len(wb.sheetnames),
        "rows": total_rows,
    }


def markdown_to_excel(input_file: str = "", input_files: list | None = None, output_file: str = "") -> dict:
    """Convert one or more Markdown files to an Excel workbook."""
    try:
        raw_inputs = []
        if input_files:
            raw_inputs.extend(input_files)
        if input_file and str(input_file).strip():
            raw_inputs.append(input_file)

        if not raw_inputs:
            return {"success": False, "error": "No input markdown files provided"}

        resolved_inputs = []
        seen_inputs = set()
        for path in raw_inputs:
            resolved = resolve_input_path(path, __file__)
            if resolved in seen_inputs:
                continue
            seen_inputs.add(resolved)
            if not os.path.exists(resolved):
                return {
                    "success": False,
                    "error": f"Input file not found: {resolved}. Please select the file using native file picker so absolute path is available."
                }
            resolved_inputs.append(resolved)

        if output_file and str(output_file).strip():
            output_file = resolve_output_file(output_file, __file__)
        elif len(resolved_inputs) == 1:
            output_file = default_single_output(resolved_inputs[0], "", ".xlsx")
        else:
            output_file = default_first_input_output(resolved_inputs, "_merged", ".xlsx", __file__)

        tables = []
        multi_input = len(resolved_inputs) > 1
        for file_index, md_path in enumerate(resolved_inputs, start=1):
            with open(md_path, "r", encoding="utf-8") as f:
                md_text = f.read()

            file_tables = _extract_markdown_tables(md_text)
            file_stem = Path(md_path).stem or f"Markdown_{file_index:02d}"

            if not file_tables:
                fallback_rows = [["Content"]]
                for line in md_text.splitlines():
                    stripped = line.strip()
                    if stripped:
                        fallback_rows.append([stripped])
                file_tables = [{"title": file_stem, "rows": fallback_rows}]

            if multi_input:
                if len(file_tables) == 1:
                    file_tables[0]["title"] = file_stem
                else:
                    for table_index, table in enumerate(file_tables, start=1):
                        table["title"] = f"{file_stem}_{table_index:02d}"

            tables.extend(file_tables)

        result = _write_excel_workbook(output_file, tables)
        result["source"] = "markdown"
        result["inputs"] = resolved_inputs
        return result
    except Exception as e:
        return {"success": False, "error": str(e)}


def pdf_to_excel(input_file: str, output_file: str = "") -> dict:
    """Convert PDF tables to an Excel workbook using pdfplumber extraction."""
    import pdfplumber

    try:
        input_file = resolve_input_path(input_file, __file__)
        if not os.path.exists(input_file):
            return {
                "success": False,
                "error": f"Input file not found: {input_file}. Please select the file using native file picker so absolute path is available."
            }

        output_file = resolve_output_file(output_file, __file__) if output_file and str(output_file).strip() else default_single_output(input_file, "", ".xlsx")

        sheets = []
        table_count = 0

        with pdfplumber.open(input_file) as pdf:
            for page_index, page in enumerate(pdf.pages, start=1):
                try:
                    tables = page.find_tables(table_settings={
                        "vertical_strategy": "lines",
                        "horizontal_strategy": "lines",
                        "snap_tolerance": 3,
                        "join_tolerance": 3,
                        "intersection_tolerance": 3,
                    })
                except Exception:
                    tables = []

                page_rows = []
                for table_index, table in enumerate(tables, start=1):
                    try:
                        rows = table.extract() or []
                    except Exception:
                        rows = []

                    normalized_rows = []
                    for row in rows:
                        values = ["" if cell is None else str(cell).strip() for cell in (row or [])]
                        if any(values):
                            normalized_rows.append(values)

                    if not normalized_rows:
                        continue

                    table_count += 1
                    sheet_title = f"Page_{page_index:03d}_Table_{table_index:02d}"
                    sheets.append({"title": sheet_title, "rows": normalized_rows})

                if sheets or page_rows:
                    continue

            if not sheets:
                with pdfplumber.open(input_file) as pdf_text:
                    fallback_rows = [["Content"]]
                    for page_index, page in enumerate(pdf_text.pages, start=1):
                        text = page.extract_text() or ""
                        lines = [line.strip() for line in text.splitlines() if line.strip()]
                        if not lines:
                            continue
                        fallback_rows.append([f"Page {page_index}"])
                        for line in lines:
                            fallback_rows.append([line])
                if len(fallback_rows) > 1:
                    sheets.append({"title": Path(input_file).stem or "PDF", "rows": fallback_rows})

        if not sheets:
            return {"success": False, "error": "No table or text content detected in PDF"}

        result = _write_excel_workbook(output_file, sheets)
        result["source"] = "pdfplumber"
        result["tables"] = table_count
        return result
    except Exception as e:
        return {"success": False, "error": str(e)}


OPERATIONS = {
    "word_to_pdf": word_to_pdf,
    "excel_to_pdf": excel_to_pdf,
    "pptx_to_pdf": pptx_to_pdf,
    "word_to_markdown": word_to_markdown,
    "excel_to_markdown": excel_to_markdown,
    "pdf_to_markdown": pdf_to_markdown,
    "images_to_pdf": images_to_pdf,
    "html_to_pdf": html_to_pdf,
    "markdown_to_pdf": markdown_to_pdf,
    "markdown_to_excel": markdown_to_excel,
    "pdf_to_excel": pdf_to_excel,
    "docx_to_txt": docx_to_txt,
    "excel_to_csv": excel_to_csv,
}

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Document Conversion Handler")
    parser.add_argument("operation", choices=list(OPERATIONS.keys()))
    parser.add_argument("params", nargs="?", help="JSON-encoded parameters")
    args = parser.parse_args()

    params = json.loads(args.params) if args.params else {}
    result = OPERATIONS[args.operation](**params)
    print(json.dumps(result, ensure_ascii=False))
